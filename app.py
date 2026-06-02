import os
import json
import math
import urllib.request
import pandas as pd
from flask import Flask, request, jsonify, render_template, send_file
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl.styles import PatternFill, Font
from datetime import datetime

app = Flask(__name__)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
EXPORT_FOLDER = os.path.join(BASE_DIR, 'exports')
CONFIG_FILE = os.path.join(BASE_DIR, 'config', 'metrics.json')

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(EXPORT_FOLDER, exist_ok=True)
os.makedirs(os.path.join(BASE_DIR, 'config'), exist_ok=True)

# ──────────────────────────────────────────────────────────────────────────────
# 常量
# ──────────────────────────────────────────────────────────────────────────────

TIME_COLUMN_KEYWORDS = ['week', 'date', 'month', 'period', 'time', 'day',
                        '周', '日期', '月份', '时间', '日']

# 固定列展示顺序：维度 → 时间 → 漏斗绝对值 → 衍生指标（由调用方追加）
FIXED_COL_ORDER = [
    'channel', 'day', 'week', 'date', 'period',
    'installs', 'register_success_events', 'first_time_events',
    'first_time_paid_events', 'cost',
]

# 指标方向标签，用于环比配色
METRIC_DIRECTION = {
    # higher_is_better
    'installs': 'higher', 'register_success_events': 'higher',
    'first_time_events': 'higher', 'first_time_paid_events': 'higher',
    '注册率': 'higher', '下单率': 'higher', '成交率': 'higher',
    # lower_is_better
    'cost': 'lower', '注册成本': 'lower', '下单成本': 'lower', '成交成本': 'lower',
}

# ──────────────────────────────────────────────────────────────────────────────
# 工具函数
# ──────────────────────────────────────────────────────────────────────────────

def detect_time_column(df):
    """自动检测时间列：先按列名关键词，再按值能否解析为日期"""
    for col in df.columns:
        if any(kw in col.lower() for kw in TIME_COLUMN_KEYWORDS):
            return col
    # 按值尝试解析
    for col in df.columns:
        sample = df[col].dropna().head(5).astype(str)
        parsed = pd.to_datetime(sample, errors='coerce', infer_datetime_format=True)
        if parsed.notna().sum() >= min(3, len(sample)):
            return col
    return None


def detect_data_mode(df, time_col):
    """
    返回 'trend'（时间趋势）或 'snapshot'（快照对比）
    - 找到时间列且有 >1 行不同时间值 → trend
    - 否则 → snapshot
    """
    if time_col and time_col in df.columns:
        unique_times = df[time_col].dropna().unique()
        if len(unique_times) > 1:
            return 'trend'
    return 'snapshot'


def sort_by_time(df, time_col):
    """按时间列升序排序，支持 YYYY/M/D 等格式"""
    if not time_col or time_col not in df.columns:
        return df
    try:
        df = df.copy()
        df['_sort_key'] = pd.to_datetime(df[time_col], errors='coerce', infer_datetime_format=True)
        df = df.sort_values('_sort_key').drop(columns=['_sort_key'])
    except Exception:
        df = df.sort_values(time_col)
    return df.reset_index(drop=True)


def ordered_display_cols(df_cols, metric_names, time_col=None):
    """
    按 FIXED_COL_ORDER 固定顺序输出列，时间列若不在固定顺序中则插到第二位。
    只保留实际存在的列。
    """
    seen = set()
    ordered = []
    base = list(FIXED_COL_ORDER)
    # 如果 time_col 不在固定顺序里，在 channel 后面插入
    if time_col and time_col not in base:
        insert_at = 1 if 'channel' in df_cols else 0
        base.insert(insert_at, time_col)
    for col in base:
        if col in df_cols and col not in seen:
            ordered.append(col)
            seen.add(col)
    # 衍生指标列追加
    for col in metric_names:
        if col in df_cols and col not in seen:
            ordered.append(col)
            seen.add(col)
    return ordered


def load_metrics_config():
    if not os.path.exists(CONFIG_FILE):
        return []
    try:
        with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (json.JSONDecodeError, IOError):
        return []


def save_metrics_config(configs):
    try:
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(configs, f, ensure_ascii=False, indent=2)
    except IOError as e:
        raise IOError(f'保存配置失败: {str(e)}')


def calculate_metrics(df, metric_defs):
    result_cols = {}
    for m in metric_defs:
        name = m['name']
        dtype = m['type']
        if m.get('isFormula') and m.get('formula'):
            raw = _eval_formula_series(df, m['formula'])
            if raw is None:
                continue
        else:
            num = m.get('numerator', '')
            den = m.get('denominator', '')
            if num not in df.columns or den not in df.columns:
                continue
            raw = df[num] / df[den].replace(0, float('nan'))
            if dtype == 'percent':
                raw = raw * 100
        result_cols[name] = pd.to_numeric(raw, errors='coerce').round(4)
    return result_cols


def _eval_formula_series(df, formula):
    import re
    fields = re.findall(r'\{([^}]+)\}', formula)
    for f in fields:
        if f not in df.columns:
            return None
    local_vars = {}
    expr = formula
    for i, f in enumerate(fields):
        var = f'col_{i}'
        expr = expr.replace('{' + f + '}', var)
        local_vars[var] = pd.to_numeric(df[f], errors='coerce')
    if not re.match(r'^[\d\s\+\-\*/\(\)\.col_0-9]+$', expr):
        return None
    try:
        return eval(expr, {"__builtins__": {}}, local_vars)  # noqa: S307
    except Exception:
        return None


def detect_anomalies(series, threshold=0.5):
    numeric = pd.to_numeric(series, errors='coerce')
    pct_change = numeric.pct_change().abs()
    return pct_change.fillna(0) > threshold


def safe_float(v):
    try:
        f = float(v)
        return None if math.isnan(f) else f
    except (TypeError, ValueError):
        return None

# ──────────────────────────────────────────────────────────────────────────────
# 路由
# ──────────────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/api/upload', methods=['POST'])
def upload_csv():
    """上传一个或多个 CSV，返回字段列表、预览数据、数据模式"""
    files = request.files.getlist('files')
    if not files:
        return jsonify({'error': '未选择文件'}), 400

    all_data = []
    for f in files:
        filename = secure_filename(f.filename)
        if not filename.endswith('.csv'):
            continue
        save_path = os.path.join(UPLOAD_FOLDER, filename)
        f.save(save_path)

        try:
            df = pd.read_csv(save_path)
        except Exception as e:
            return jsonify({'error': f'解析 {filename} 失败: {str(e)}'}), 400

        time_col = detect_time_column(df)
        mode = detect_data_mode(df, time_col)

        app_values = sorted(df['app'].dropna().unique().tolist()) if 'app' in df.columns else []
        date_values = []
        if time_col and time_col in df.columns:
            date_values = sorted(df[time_col].dropna().astype(str).unique().tolist())

        all_data.append({
            'filename': filename,
            'platform': filename.replace('.csv', ''),
            'columns': list(df.columns),
            'time_column': time_col,
            'mode': mode,
            'rows': len(df),
            'preview': df.head(3).fillna('').to_dict(orient='records'),
            'app_values': app_values,
            'date_values': date_values,
        })

    if not all_data:
        return jsonify({'error': '未找到有效的 CSV 文件'}), 400

    return jsonify({'files': all_data})


@app.route('/api/analyze', methods=['POST'])
def analyze():
    """
    执行分析。
    根据 mode 字段（snapshot / trend）走不同分支：
      snapshot → 环比对比（最多2个文件，按 channel 对齐）
      trend    → 时间趋势（先升序排序，折线图）
    """
    payload = request.get_json()
    files_info   = payload.get('files', [])
    metric_defs  = payload.get('metrics', [])
    filter_apps  = payload.get('filter_apps', [])
    filter_date_start = payload.get('filter_date_start', '')
    filter_date_end   = payload.get('filter_date_end', '')
    # 前端可手动覆盖模式：'snapshot' / 'trend' / 'auto'
    mode_override = payload.get('mode', 'auto')

    if not files_info:
        return jsonify({'error': '请先上传 CSV 文件'}), 400

    # 读取并预处理所有 df
    dfs = []
    for fi in files_info:
        filename = fi['filename']
        platform = fi.get('platform', filename)
        time_col = fi.get('time_column')

        csv_path = os.path.join(UPLOAD_FOLDER, secure_filename(filename))
        if not os.path.exists(csv_path):
            continue
        try:
            df = pd.read_csv(csv_path)
        except Exception:
            continue

        if filter_apps and 'app' in df.columns:
            df = df[df['app'].astype(str).isin(filter_apps)]
        if time_col and time_col in df.columns:
            if filter_date_start:
                df = df[df[time_col].astype(str) >= filter_date_start]
            if filter_date_end:
                df = df[df[time_col].astype(str) <= filter_date_end]
        if df.empty:
            continue

        dfs.append({'df': df, 'platform': platform, 'time_col': time_col,
                    'filename': filename})

    if not dfs:
        return jsonify({'error': '筛选后数据为空'}), 400

    # 确定最终模式
    if mode_override == 'auto':
        # 用第一个文件的 time_col 自动判断
        first_tc = dfs[0]['time_col']
        mode = detect_data_mode(dfs[0]['df'], first_tc)
    else:
        mode = mode_override

    if mode == 'snapshot':
        return _analyze_snapshot(dfs, metric_defs)
    else:
        return _analyze_trend(dfs, metric_defs)


# ─────────────────────────────────────────────────────────────
# 模式A：快照对比 / 周环比
# ─────────────────────────────────────────────────────────────

def _analyze_snapshot(dfs, metric_defs):
    """
    按 channel 对齐两个文件，计算环比。
    返回结构与 trend 不同，供前端按 mode 分支渲染。
    wow_rows: [{metric, current_val, base_val, delta, delta_pct, direction}]
    charts:   {metric: {type, labels, datasets:[{label,data}]}}
    tables:   [{platform, columns, rows, anomalies}]
    """
    metric_names = [m['name'] for m in metric_defs]
    direction_map = {**METRIC_DIRECTION}
    for m in metric_defs:
        if m['name'] not in direction_map:
            direction_map[m['name']] = 'higher'

    # 计算每个文件的指标
    processed = []
    for item in dfs:
        df = item['df'].copy()
        extra = calculate_metrics(df, metric_defs)
        for col, series in extra.items():
            df[col] = series
        processed.append({**item, 'df': df})

    # 整理展示用表格
    tables_data = []
    for item in processed:
        df = item['df']
        m_names = [m['name'] for m in metric_defs if m['name'] in df.columns]
        cols = ordered_display_cols(list(df.columns), m_names, item['time_col'])
        table_rows = df[cols].fillna('').to_dict(orient='records')
        anomalies = {}
        for col in m_names:
            anom = detect_anomalies(df[col])
            anomalies[col] = [i for i, v in enumerate(anom) if v]
        tables_data.append({
            'platform': item['platform'],
            'columns': cols,
            'rows': table_rows,
            'anomalies': anomalies,
        })

    # 如果只有一个文件，返回普通快照（无环比）
    if len(processed) < 2:
        charts_data = _build_bar_charts(processed, metric_defs)
        return jsonify({'mode': 'snapshot', 'charts': charts_data,
                        'tables': tables_data, 'wow': None})

    # 两个文件：本期=processed[0]，基期=processed[1]
    cur_item = processed[0]
    base_item = processed[1]
    cur_df  = cur_item['df']
    base_df = base_item['df']

    # 确定对齐键
    all_cols = set(cur_df.columns) & set(base_df.columns)
    align_col = 'channel' if 'channel' in all_cols else None

    wow_rows = []
    abs_cols = ['installs', 'register_success_events', 'first_time_events',
                'first_time_paid_events', 'cost']
    compare_cols = [c for c in abs_cols if c in cur_df.columns and c in base_df.columns] + \
                   [m['name'] for m in metric_defs
                    if m['name'] in cur_df.columns and m['name'] in base_df.columns]

    if align_col:
        # 按 channel 聚合（若同一文件有多行相同 channel，先 groupby sum/mean）
        def agg_df(df):
            num_cols = [c for c in compare_cols if c in df.columns]
            # 转化率类用 mean，绝对值类用 sum
            agg = {}
            for c in num_cols:
                m_def = next((m for m in metric_defs if m['name'] == c), None)
                agg[c] = 'mean' if (m_def and m_def['type'] == 'percent') else 'sum'
            if num_cols:
                return df.groupby(align_col, as_index=True)[num_cols].agg(agg)
            return df.set_index(align_col)[[]]

        cur_agg  = agg_df(cur_df)
        base_agg = agg_df(base_df)
        channels = sorted(set(cur_agg.index) & set(base_agg.index))
        if not channels:
            return jsonify({'mode': 'snapshot',
                            'error': '渠道不一致，无法对比',
                            'charts': {}, 'tables': tables_data, 'wow': None})

        # 按 channel 汇总后计算环比（跨渠道求和/均值后整体对比）
        cur_vals  = {}
        base_vals = {}
        for col in compare_cols:
            if col not in cur_agg.columns or col not in base_agg.columns:
                continue
            m_def = next((m for m in metric_defs if m['name'] == col), None)
            use_mean = m_def and m_def['type'] == 'percent'
            c_arr = cur_agg.loc[channels, col].dropna()
            b_arr = base_agg.loc[channels, col].dropna()
            cur_vals[col]  = float(c_arr.mean() if use_mean else c_arr.sum())
            base_vals[col] = float(b_arr.mean() if use_mean else b_arr.sum())
    else:
        # 无 channel 列，直接对整个文件求和/均值
        for col in compare_cols:
            if col not in cur_df.columns or col not in base_df.columns:
                continue
            m_def = next((m for m in metric_defs if m['name'] == col), None)
            use_mean = m_def and m_def['type'] == 'percent'
            cur_vals  = {col: float(cur_df[col].mean() if use_mean else cur_df[col].sum())
                         for col in compare_cols if col in cur_df.columns}
            base_vals = {col: float(base_df[col].mean() if use_mean else base_df[col].sum())
                         for col in compare_cols if col in base_df.columns}
            break  # 统一处理，不需要循环

    for col in compare_cols:
        cv = cur_vals.get(col)
        bv = base_vals.get(col)
        if cv is None or bv is None:
            continue
        delta = cv - bv
        delta_pct = (delta / bv * 100) if bv != 0 else None
        direction = direction_map.get(col, 'higher')
        wow_rows.append({
            'metric': col,
            'current': round(cv, 4),
            'base': round(bv, 4),
            'delta': round(delta, 4),
            'delta_pct': round(delta_pct, 2) if delta_pct is not None else None,
            'direction': direction,
        })

    charts_data = _build_bar_charts(processed, metric_defs)
    return jsonify({
        'mode': 'snapshot',
        'charts': charts_data,
        'tables': tables_data,
        'wow': {
            'current_label': cur_item['platform'],
            'base_label': base_item['platform'],
            'rows': wow_rows,
        },
    })


def _build_bar_charts(processed, metric_defs):
    """为快照模式构建分组柱状图数据（每文件一个数据点，按 metric 分组）"""
    charts_data = {}
    for item in processed:
        df = item['df']
        platform = item['platform']
        for m in metric_defs:
            mname = m['name']
            if mname not in df.columns:
                continue
            m_def_type = m.get('type', 'count')
            use_mean = m_def_type == 'percent'
            val = float(df[mname].mean() if use_mean else df[mname].sum())
            if math.isnan(val):
                continue
            if mname not in charts_data:
                charts_data[mname] = {
                    'labels': [],
                    'datasets': [{'label': mname, 'data': []}],
                    'type': m_def_type,
                    'currency': m.get('currency', ''),
                    'unit': m.get('unit', ''),
                    'chartType': 'bar',
                }
            charts_data[mname]['labels'].append(platform)
            charts_data[mname]['datasets'][0]['data'].append(round(val, 4))
    return charts_data


# ─────────────────────────────────────────────────────────────
# 模式B：时间趋势
# ─────────────────────────────────────────────────────────────

def _analyze_trend(dfs, metric_defs):
    """
    时间趋势模式：先排序，再按时间轴对齐多文件。
    """
    charts_data = {}
    tables_data = []

    # 收集所有时间轴标签（并集，升序）
    all_labels_set = set()
    for item in dfs:
        df = item['df']
        tc = item['time_col']
        if tc and tc in df.columns:
            all_labels_set.update(df[tc].dropna().astype(str).tolist())

    # 尝试日期排序
    def label_sort_key(s):
        try:
            return pd.to_datetime(s, infer_datetime_format=True)
        except Exception:
            return s
    all_labels = sorted(all_labels_set, key=label_sort_key)

    for item in dfs:
        df = item['df'].copy()
        platform = item['platform']
        time_col = item['time_col']

        # 升序排序
        df = sort_by_time(df, time_col)

        # 计算衍生指标
        extra = calculate_metrics(df, metric_defs)
        for col, series in extra.items():
            df[col] = series

        # 构建时间→行的映射（用于对齐全局 labels）
        if time_col and time_col in df.columns:
            row_map = {str(row[time_col]): row for _, row in df.iterrows()}
        else:
            row_map = {str(i): row for i, (_, row) in enumerate(df.iterrows())}

        # 图表数据（时间轴对齐，缺失填 None）
        for m in metric_defs:
            mname = m['name']
            data_points = []
            for lbl in all_labels:
                row = row_map.get(lbl)
                if row is None:
                    data_points.append(None)
                else:
                    v = safe_float(row.get(mname))
                    data_points.append(round(v, 4) if v is not None else None)

            if all(p is None for p in data_points):
                continue

            if mname not in charts_data:
                charts_data[mname] = {
                    'labels': all_labels,
                    'datasets': [],
                    'type': m['type'],
                    'currency': m.get('currency', ''),
                    'unit': m.get('unit', ''),
                    'chartType': 'line',
                }
            charts_data[mname]['datasets'].append({
                'label': platform,
                'data': data_points,
            })

        # 表格数据（只用自己的行，已排序）
        m_names = [m['name'] for m in metric_defs if m['name'] in df.columns]
        cols = ordered_display_cols(list(df.columns), m_names, time_col)
        table_rows = df[cols].fillna('').to_dict(orient='records')
        anomalies = {}
        for col in m_names:
            anom = detect_anomalies(df[col])
            anomalies[col] = [i for i, v in enumerate(anom) if v]
        tables_data.append({
            'platform': platform,
            'columns': cols,
            'rows': table_rows,
            'anomalies': anomalies,
        })

    return jsonify({
        'mode': 'trend',
        'charts': charts_data,
        'tables': tables_data,
        'wow': None,
    })


# ──────────────────────────────────────────────────────────────────────────────
# 指标配置 CRUD
# ──────────────────────────────────────────────────────────────────────────────

@app.route('/api/metrics', methods=['GET'])
def get_metrics_configs():
    return jsonify(load_metrics_config())

@app.route('/api/metrics', methods=['POST'])
def save_metrics():
    data = request.get_json()
    name = data.get('name', '').strip()
    metrics = data.get('metrics', [])
    if not name:
        return jsonify({'error': '配置名称不能为空'}), 400
    configs = load_metrics_config()
    configs = [c for c in configs if c['name'] != name]
    configs.append({'name': name, 'metrics': metrics})
    save_metrics_config(configs)
    return jsonify({'success': True})

@app.route('/api/metrics/<name>', methods=['DELETE'])
def delete_metrics(name):
    configs = load_metrics_config()
    configs = [c for c in configs if c['name'] != name]
    save_metrics_config(configs)
    return jsonify({'success': True})


# ──────────────────────────────────────────────────────────────────────────────
# Excel 导出
# ──────────────────────────────────────────────────────────────────────────────

@app.route('/api/export/excel', methods=['POST'])
def export_excel():
    payload = request.get_json()
    tables_data = payload.get('tables', [])
    metric_defs = payload.get('metrics', [])
    config_name = payload.get('config_name', 'export')

    dtype_map = {m['name']: m for m in metric_defs}
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_fill  = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font  = Font(color="FFFFFF", bold=True)
    anomaly_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")

    used_sheet_names = set()
    for table in tables_data:
        platform_raw = table['platform'][:31]
        platform = platform_raw
        suffix = 1
        while platform in used_sheet_names:
            platform = f"{platform_raw[:28]}_{suffix}"
            suffix += 1
        used_sheet_names.add(platform)
        ws = wb.create_sheet(title=platform)

        columns  = table['columns']
        rows     = table['rows']
        anomalies = table.get('anomalies', {})

        for ci, col in enumerate(columns, 1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.fill = header_fill
            cell.font = header_font

        for ri, row in enumerate(rows, 2):
            for ci, col in enumerate(columns, 1):
                val = row.get(col, '')
                if col in dtype_map:
                    m = dtype_map[col]
                    try:
                        numeric_val = float(val)
                        cell = ws.cell(row=ri, column=ci, value=numeric_val)
                        if m['type'] == 'percent':
                            cell.number_format = '0.00"%"'
                        elif m['type'] == 'currency':
                            sym = m.get('currency', '')
                            cell.number_format = f'"{sym}"#,##0.00'
                        elif m['type'] == 'count':
                            cell.number_format = '#,##0'
                    except (ValueError, TypeError):
                        ws.cell(row=ri, column=ci, value=val)
                else:
                    ws.cell(row=ri, column=ci, value=val)
                if col in anomalies and (ri - 2) in anomalies[col]:
                    ws.cell(row=ri, column=ci).fill = anomaly_fill

        for col_cells in ws.columns:
            max_len = max((len(str(c.value)) for c in col_cells if c.value), default=10)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 30)

    timestamp  = datetime.now().strftime('%Y%m%d_%H%M%S')
    safe_name  = config_name.replace('/', '_').replace('\\', '_')
    filename   = f"{timestamp}_{safe_name}.xlsx"
    export_path = os.path.join(EXPORT_FOLDER, filename)
    try:
        wb.save(export_path)
    except Exception as e:
        return jsonify({'error': f'Excel 导出失败: {str(e)}'}), 500

    return send_file(export_path, as_attachment=True, download_name=filename)


# ──────────────────────────────────────────────────────────────────────────────
# AI 分析代理
# ──────────────────────────────────────────────────────────────────────────────

@app.route('/api/ai-analyze', methods=['POST'])
def ai_analyze():
    payload = request.get_json()
    api_key      = payload.get('api_key', '').strip()
    data_summary = payload.get('data_summary', '')

    if not api_key:
        return jsonify({'error': '请填写 Anthropic API Key'}), 400
    if not data_summary:
        return jsonify({'error': '数据摘要为空'}), 400

    system_prompt = (
        "你是一位拥有10年经验的资深投放数据优化师，专注于效果广告投放优化，"
        "熟悉各大流量平台（字节、腾讯、百度、Meta、Google等）的算法机制和优化策略。\n"
        "根据用户提供的投放数据，给出专业、深度、可落地的分析报告。\n\n"
        "报告结构（严格按此格式输出）：\n"
        "## 一、数据核心发现\n列出3-5个关键数据洞察，每条需有具体数值支撑\n\n"
        "## 二、问题诊断\n识别表现异常的指标，分析可能原因\n\n"
        "## 三、优化建议\n给出5-8条具体可执行建议，每条标注优先级：🔴紧急 / 🟡重要 / 🟢建议\n\n"
        "## 四、平台对比结论\n（如有多平台数据）指出哪个平台表现更优及原因；单平台则分析趋势\n\n"
        "## 五、下阶段策略\n给出简明的下阶段投放策略方向\n\n"
        "语言要求：专业但易懂，用数据说话，避免空话套话。"
    )

    body = json.dumps({
        'model': 'claude-sonnet-4-5',
        'max_tokens': 2048,
        'system': system_prompt,
        'messages': [{'role': 'user', 'content': f'请分析以下投放数据并给出专业建议：\n\n{data_summary}'}]
    }).encode('utf-8')

    req = urllib.request.Request(
        'https://api.anthropic.com/v1/messages',
        data=body,
        headers={
            'Content-Type': 'application/json',
            'x-api-key': api_key,
            'anthropic-version': '2023-06-01',
        },
        method='POST'
    )

    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            result = json.loads(resp.read().decode('utf-8'))
        text = ''.join(b.get('text', '') for b in result.get('content', []))
        return jsonify({'result': text})
    except urllib.error.HTTPError as e:
        err_body = e.read().decode('utf-8', errors='replace')
        try:
            err_json = json.loads(err_body)
            msg = err_json.get('error', {}).get('message', err_body)
        except Exception:
            msg = err_body
        return jsonify({'error': f'API 错误 {e.code}: {msg}'}), 502
    except Exception as e:
        return jsonify({'error': f'请求失败: {str(e)}'}), 502


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5001))
    print("=" * 50)
    print("投放数据分析工具已启动")
    print(f"请在浏览器访问: http://127.0.0.1:{port}")
    print("=" * 50)
    app.run(debug=False, port=port, host='0.0.0.0')
